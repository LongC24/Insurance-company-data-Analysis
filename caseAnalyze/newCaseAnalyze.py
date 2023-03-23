from datetime import datetime, timedelta
import prettytable as pt
from openpyxl.utils import get_column_letter

import connectDB as cdb
import datesGenerate
from openpyxl import Workbook


# dates = generate_dates('2022-01-01', '2023-03-28')

def pt_single_date_print(date):
    date = cdb.single_submit_date('2023-02-09')
    # 用 prettytable 打印出来
    table = pt.PrettyTable()
    table.field_names = cdb.Column
    for i in range(len(date)):
        table.add_row(date[i])
    print(table)


def pt_range_date_print(start_date, end_date, debug=False):
    monday, sunday = datesGenerate.get_mondays_and_sundays(start_date, end_date)
    if debug:
        print(monday)
        print(sunday)

    # 便利日期列表中的日期 从数据库中获取这一周的日期 并打印出来
    for i in range(len(monday)):
        print()
        print(monday[i] + '  ********************************  ' + sunday[i])
        date = cdb.range_submit_date(monday[i], sunday[i])
        table = pt.PrettyTable()
        table.field_names = cdb.Column
        for i in range(len(date)):
            table.add_row(date[i])
        print(table)


def write_excel_weeks_with_data(start_date, end_date, file_name='WeekReport', file_path='./exportFile/', debug=False):
    wb = Workbook()
    # 获取所有的周一和周日
    monday, sunday = datesGenerate.get_mondays_and_sundays(start_date, end_date)
    if debug:
        print(monday)
        print(sunday)
    # 便利日期列表中的日期 从数据库中获取这一周的日期 并添加到表格中
    for i in range(len(monday)):
        # 创建工作表名字
        sheet_name = str(monday[i] + '~' + sunday[i])
        ws = wb.create_sheet(sheet_name)
        ws.title = sheet_name

        # 获取这一周的数据
        date = cdb.range_submit_date(monday[i], sunday[i])
        if debug:
            print()
            print(monday[i] + '  ********************************  ' + sunday[i])
            print(date)

        # 在工作表中第一行写入这一周有多少人交单（既有多少行）
        weeks_count = "本周交单数量： " + str(len(date))
        ws.cell(row=1, column=1, value=weeks_count)
        # 把第一行 前10个格子合并成一个
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        # 在工作表中第二行写入列名
        for i in range(len(cdb.Column)):
            ws.cell(row=2, column=i + 1, value=cdb.Column[i])
        # 写入数据
        for i in range(len(date)):
            for j in range(len(date[i])):
                ws.cell(row=i + 3, column=j + 1, value=date[i][j])
    # 删除第一个表
    wb.remove(wb['Sheet'])

    # 设置所有的表格 为 自动宽度
    for sheet in wb:
        for column_cells in sheet.columns:
            # 计算该列中最长单元格的长度
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            # 获取该列字母
            new_column_letter = (get_column_letter(column_cells[0].column))
            # 如果该列有内容
            if new_column_length > 0:
                # 设置该列宽度为最长单元格长度乘以1.23
                sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23

    # 保存文件
    wb.save(file_path + file_name + ".xlsx")


if __name__ == '__main__':
    # pt_single_date_print('2023-02-09')
    # pt_range_date_print('2022-01-01', '2023-03-28')
    write_excel_weeks_with_data('2022-03-01', '2023-04-01', file_name='WeekReport', file_path='./exportFile/',
                                debug=True)
